"""Chat list parser for various file formats."""

from __future__ import annotations

import csv
import io
import re
from enum import Enum
from typing import BinaryIO

from pydantic import BaseModel, ConfigDict, field_validator


class ParseError(Exception):
    """Error during chat list parsing."""


class ChatListEntryType(str, Enum):
    """Type of chat list entry."""

    USERNAME = "username"  # @channel_name or just channel_name
    LINK = "link"  # https://t.me/channel_name or t.me/channel_name
    ID = "id"  # Numeric chat ID


class ChatListEntry(BaseModel):
    """A single entry from imported chat list.

    Attributes:
        value: The raw value (username, link, or ID).
        entry_type: Type of entry (username, link, or id).
        normalized: Normalized value for lookup (username without @, resolved from link).
    """

    model_config = ConfigDict(
        strict=True,
        frozen=True,
        extra="forbid",
    )

    value: str
    entry_type: ChatListEntryType
    normalized: str

    @field_validator("value")
    @classmethod
    def value_not_empty(cls, v: str) -> str:
        """Validate that value is not empty."""
        if not v.strip():
            raise ValueError("value cannot be empty")
        return v.strip()


# Patterns for parsing
TELEGRAM_LINK_PATTERN = re.compile(
    r"(?:https?://)?(?:t\.me|telegram\.me)/(?:joinchat/|\+)?([a-zA-Z0-9_-]+)",
    re.IGNORECASE,
)
USERNAME_PATTERN = re.compile(r"^@?([a-zA-Z][a-zA-Z0-9_]{3,31})$")
NUMERIC_ID_PATTERN = re.compile(r"^-?\d+$")

# Dangerous characters that could break CSV, SSE, or DB operations
# Control chars: \x00 (null), \x01-\x1f (control), \x7f-\x9f (extended control)
# Notably includes: \n (0x0a), \r (0x0d), \t (0x09)
DANGEROUS_CHARS_PATTERN = re.compile(r"[\x00-\x1f\x7f-\x9f]")


def _validate_chat_ref_safety(value: str) -> None:
    """Validate that chat_ref doesn't contain dangerous characters.

    Args:
        value: Chat reference string to validate.

    Raises:
        ParseError: If the value contains dangerous characters.
    """
    # Check for control characters (null bytes, newlines, etc.)
    if DANGEROUS_CHARS_PATTERN.search(value):
        # Identify specific issue for better error message
        if "\x00" in value:
            raise ParseError(f"Invalid chat reference: contains null byte: {value[:50]!r}")
        elif "\n" in value or "\r" in value:
            raise ParseError(f"Invalid chat reference: contains newline: {value[:50]!r}")
        elif "\t" in value:
            raise ParseError(f"Invalid chat reference: contains tab character: {value[:50]!r}")
        else:
            # Other control characters
            raise ParseError(f"Invalid chat reference: contains control characters: {value[:50]!r}")

    # Must be valid UTF-8
    try:
        value.encode("utf-8")
    except UnicodeEncodeError as e:
        raise ParseError(f"Invalid chat reference: invalid UTF-8 encoding: {value[:50]!r}") from e


def _classify_entry(raw: str) -> ChatListEntry | None:
    """Classify a single entry and create ChatListEntry.

    Args:
        raw: Raw string value from the file.

    Returns:
        ChatListEntry if valid, None if the line should be skipped.

    Raises:
        ParseError: If the value contains dangerous characters.
    """
    value = raw.strip()

    # Skip empty lines and comments
    if not value or value.startswith("#"):
        return None

    # Validate for dangerous characters BEFORE processing
    # This will raise ParseError if invalid, stopping the entire import
    _validate_chat_ref_safety(value)

    # Check for Telegram link
    link_match = TELEGRAM_LINK_PATTERN.match(value)
    if link_match:
        username = link_match.group(1)
        return ChatListEntry(
            value=value,
            entry_type=ChatListEntryType.LINK,
            normalized=username.lower(),
        )

    # Check for numeric ID
    if NUMERIC_ID_PATTERN.match(value):
        return ChatListEntry(
            value=value,
            entry_type=ChatListEntryType.ID,
            normalized=value,
        )

    # Check for username (with or without @)
    username_match = USERNAME_PATTERN.match(value)
    if username_match:
        username = username_match.group(1)
        return ChatListEntry(
            value=value,
            entry_type=ChatListEntryType.USERNAME,
            normalized=username.lower(),
        )

    # If it looks like a username but doesn't match pattern, still try
    # (could be a display name or partial match)
    clean = value.lstrip("@").strip()
    if clean and len(clean) >= 2:
        return ChatListEntry(
            value=value,
            entry_type=ChatListEntryType.USERNAME,
            normalized=clean.lower(),
        )

    return None


def parse_text(content: str | bytes) -> list[ChatListEntry]:
    """Parse a text file with one chat per line.

    Args:
        content: File content as string or bytes.

    Returns:
        List of parsed chat entries.

    Raises:
        ParseError: If parsing fails.
    """
    text: str
    if isinstance(content, bytes):
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = content.decode("cp1251")  # Common for Russian text
            except UnicodeDecodeError as e:
                raise ParseError(f"Failed to decode file: {e}") from e
    else:
        text = content

    entries = []
    for line in text.splitlines():
        entry = _classify_entry(line)
        if entry:
            entries.append(entry)

    return entries


def parse_csv(content: str | bytes) -> list[ChatListEntry]:
    """Parse a CSV file with chat data.

    Looks for columns named: username, chat, link, id, url, name
    Falls back to first column if no recognized header found.

    Args:
        content: File content as string or bytes.

    Returns:
        List of parsed chat entries.

    Raises:
        ParseError: If parsing fails.
    """
    text: str
    if isinstance(content, bytes):
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = content.decode("cp1251")
            except UnicodeDecodeError as e:
                raise ParseError(f"Failed to decode CSV: {e}") from e
    else:
        text = content

    entries = []
    try:
        # Detect delimiter
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        # Fallback to comma
        reader = csv.reader(io.StringIO(text))

    rows = list(reader)
    if not rows:
        return []

    # Check for header row
    header = rows[0]
    header_lower = [h.lower().strip() for h in header]

    # Known column names for chat identifiers
    chat_columns = {"username", "chat", "link", "id", "url", "name", "channel", "group"}
    col_index = None

    for idx, col_name in enumerate(header_lower):
        if col_name in chat_columns:
            col_index = idx
            break

    # If no recognized header, use first column and include first row
    if col_index is None:
        col_index = 0
        # Check if first row looks like data (not a header)
        first_entry = _classify_entry(rows[0][0] if rows[0] else "")
        start_row = 0 if first_entry else 1
    else:
        start_row = 1  # Skip header

    for row in rows[start_row:]:
        if col_index < len(row):
            entry = _classify_entry(row[col_index])
            if entry:
                entries.append(entry)

    return entries


def parse_xlsx(file: BinaryIO) -> list[ChatListEntry]:
    """Parse an Excel file with chat data.

    Looks for columns named: username, chat, link, id, url, name
    Falls back to first column if no recognized header found.

    Args:
        file: File-like object with Excel content.

    Returns:
        List of parsed chat entries.

    Raises:
        ParseError: If parsing fails.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ParseError("openpyxl is required for Excel file support") from e

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"Failed to read Excel file: {e}") from e

    entries = []
    sheet = wb.active

    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Check for header row
    header = [str(cell or "").lower().strip() for cell in rows[0]]

    # Known column names
    chat_columns = {"username", "chat", "link", "id", "url", "name", "channel", "group"}
    col_index = None

    for idx, col_name in enumerate(header):
        if col_name in chat_columns:
            col_index = idx
            break

    # If no recognized header, use first column
    if col_index is None:
        col_index = 0
        # Check if first row looks like data
        first_val = str(rows[0][0] or "") if rows[0] else ""
        first_entry = _classify_entry(first_val)
        start_row = 0 if first_entry else 1
    else:
        start_row = 1

    for row in rows[start_row:]:
        if col_index < len(row) and row[col_index]:
            entry = _classify_entry(str(row[col_index]))
            if entry:
                entries.append(entry)

    wb.close()
    return entries


def parse_chat_list(
    content: bytes,
    filename: str,
) -> list[ChatListEntry]:
    """Parse a chat list file based on its extension.

    Args:
        content: File content as bytes.
        filename: Original filename (used to detect format).

    Returns:
        List of parsed chat entries.

    Raises:
        ParseError: If parsing fails or format is unsupported.
    """
    filename_lower = filename.lower()

    if filename_lower.endswith(".txt"):
        return parse_text(content)
    elif filename_lower.endswith(".csv"):
        return parse_csv(content)
    elif filename_lower.endswith((".xlsx", ".xls")):
        return parse_xlsx(io.BytesIO(content))
    else:
        # Try to detect format
        # If it starts with PK, it's likely a zip-based format (xlsx)
        if content.startswith(b"PK"):
            return parse_xlsx(io.BytesIO(content))
        # Try CSV first, then plain text
        try:
            return parse_csv(content)
        except ParseError:
            return parse_text(content)
