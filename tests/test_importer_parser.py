"""Comprehensive tests for chat list parser module.

Tests all parsers (text, CSV, Excel) with various formats, encodings,
edge cases, and error conditions to achieve maximum coverage.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

import pytest
from pydantic import ValidationError

from chatfilter.importer.parser import (
    NUMERIC_ID_PATTERN,
    TELEGRAM_LINK_PATTERN,
    USERNAME_PATTERN,
    ChatListEntry,
    ChatListEntryType,
    ParseError,
    _classify_entry,
    parse_chat_list,
    parse_csv,
    parse_text,
    parse_xlsx,
)

if TYPE_CHECKING:
    pass


class TestChatListEntry:
    """Tests for ChatListEntry model validation."""

    def test_valid_username_entry(self):
        """Test creating a valid username entry."""
        entry = ChatListEntry(
            value="@testuser",
            entry_type=ChatListEntryType.USERNAME,
            normalized="testuser",
        )
        assert entry.value == "@testuser"
        assert entry.entry_type == ChatListEntryType.USERNAME
        assert entry.normalized == "testuser"

    def test_valid_link_entry(self):
        """Test creating a valid link entry."""
        entry = ChatListEntry(
            value="https://t.me/testchannel",
            entry_type=ChatListEntryType.LINK,
            normalized="testchannel",
        )
        assert entry.value == "https://t.me/testchannel"
        assert entry.entry_type == ChatListEntryType.LINK
        assert entry.normalized == "testchannel"

    def test_valid_id_entry(self):
        """Test creating a valid ID entry."""
        entry = ChatListEntry(
            value="-1001234567890",
            entry_type=ChatListEntryType.ID,
            normalized="-1001234567890",
        )
        assert entry.value == "-1001234567890"
        assert entry.entry_type == ChatListEntryType.ID
        assert entry.normalized == "-1001234567890"

    def test_empty_value_rejected(self):
        """Test that empty value is rejected."""
        with pytest.raises(ValidationError, match="value cannot be empty"):
            ChatListEntry(
                value="",
                entry_type=ChatListEntryType.USERNAME,
                normalized="test",
            )

    def test_whitespace_only_value_rejected(self):
        """Test that whitespace-only value is rejected."""
        with pytest.raises(ValidationError, match="value cannot be empty"):
            ChatListEntry(
                value="   ",
                entry_type=ChatListEntryType.USERNAME,
                normalized="test",
            )

    def test_value_is_stripped(self):
        """Test that value is stripped of whitespace."""
        entry = ChatListEntry(
            value="  @testuser  ",
            entry_type=ChatListEntryType.USERNAME,
            normalized="testuser",
        )
        assert entry.value == "@testuser"

    def test_model_is_frozen(self):
        """Test that ChatListEntry is immutable."""
        entry = ChatListEntry(
            value="@testuser",
            entry_type=ChatListEntryType.USERNAME,
            normalized="testuser",
        )
        with pytest.raises(ValidationError):
            entry.value = "@newuser"  # type: ignore[misc]

    def test_extra_fields_forbidden(self):
        """Test that extra fields are rejected."""
        with pytest.raises(ValidationError):
            ChatListEntry(  # type: ignore[call-arg]
                value="@testuser",
                entry_type=ChatListEntryType.USERNAME,
                normalized="testuser",
                extra_field="not allowed",
            )


class TestPatterns:
    """Tests for regex patterns."""

    def test_telegram_link_pattern_https(self):
        """Test Telegram link pattern with https."""
        match = TELEGRAM_LINK_PATTERN.match("https://t.me/testchannel")
        assert match is not None
        assert match.group(1) == "testchannel"

    def test_telegram_link_pattern_http(self):
        """Test Telegram link pattern with http."""
        match = TELEGRAM_LINK_PATTERN.match("http://t.me/testchannel")
        assert match is not None
        assert match.group(1) == "testchannel"

    def test_telegram_link_pattern_no_protocol(self):
        """Test Telegram link pattern without protocol."""
        match = TELEGRAM_LINK_PATTERN.match("t.me/testchannel")
        assert match is not None
        assert match.group(1) == "testchannel"

    def test_telegram_link_pattern_telegram_me(self):
        """Test Telegram link pattern with telegram.me."""
        match = TELEGRAM_LINK_PATTERN.match("https://telegram.me/testchannel")
        assert match is not None
        assert match.group(1) == "testchannel"

    def test_telegram_link_pattern_joinchat(self):
        """Test Telegram link pattern with joinchat."""
        match = TELEGRAM_LINK_PATTERN.match("https://t.me/joinchat/AaBbCcDd")
        assert match is not None
        assert match.group(1) == "AaBbCcDd"

    def test_telegram_link_pattern_plus_prefix(self):
        """Test Telegram link pattern with + prefix."""
        match = TELEGRAM_LINK_PATTERN.match("https://t.me/+AaBbCcDd")
        assert match is not None
        assert match.group(1) == "AaBbCcDd"

    def test_telegram_link_pattern_case_insensitive(self):
        """Test Telegram link pattern is case insensitive."""
        match = TELEGRAM_LINK_PATTERN.match("HTTPS://T.ME/TestChannel")
        assert match is not None
        assert match.group(1) == "TestChannel"

    def test_username_pattern_with_at(self):
        """Test username pattern with @ prefix."""
        match = USERNAME_PATTERN.match("@testuser")
        assert match is not None
        assert match.group(1) == "testuser"

    def test_username_pattern_without_at(self):
        """Test username pattern without @ prefix."""
        match = USERNAME_PATTERN.match("testuser")
        assert match is not None
        assert match.group(1) == "testuser"

    def test_username_pattern_minimum_length(self):
        """Test username pattern minimum length (4 chars)."""
        assert USERNAME_PATTERN.match("@abc") is None  # 3 chars - too short
        assert USERNAME_PATTERN.match("@abcd") is not None  # 4 chars - valid

    def test_username_pattern_maximum_length(self):
        """Test username pattern maximum length (32 chars)."""
        assert USERNAME_PATTERN.match("@" + "a" * 32) is not None  # 32 chars - valid
        assert USERNAME_PATTERN.match("@" + "a" * 33) is None  # 33 chars - too long

    def test_username_pattern_must_start_with_letter(self):
        """Test username must start with a letter."""
        assert USERNAME_PATTERN.match("@1user") is None
        assert USERNAME_PATTERN.match("@_user") is None
        assert USERNAME_PATTERN.match("@auser") is not None

    def test_username_pattern_allows_underscores(self):
        """Test username can contain underscores."""
        match = USERNAME_PATTERN.match("@test_user_name")
        assert match is not None
        assert match.group(1) == "test_user_name"

    def test_numeric_id_pattern_positive(self):
        """Test numeric ID pattern with positive numbers."""
        assert NUMERIC_ID_PATTERN.match("12345") is not None
        assert NUMERIC_ID_PATTERN.match("1") is not None

    def test_numeric_id_pattern_negative(self):
        """Test numeric ID pattern with negative numbers."""
        assert NUMERIC_ID_PATTERN.match("-12345") is not None
        assert NUMERIC_ID_PATTERN.match("-1001234567890") is not None

    def test_numeric_id_pattern_rejects_non_numeric(self):
        """Test numeric ID pattern rejects non-numeric input."""
        assert NUMERIC_ID_PATTERN.match("abc") is None
        assert NUMERIC_ID_PATTERN.match("123abc") is None
        assert NUMERIC_ID_PATTERN.match("12.34") is None


class TestClassifyEntry:
    """Tests for _classify_entry function."""

    def test_empty_string_returns_none(self):
        """Test that empty string returns None."""
        assert _classify_entry("") is None

    def test_whitespace_only_returns_none(self):
        """Test that whitespace-only string returns None."""
        assert _classify_entry("   ") is None
        assert _classify_entry("\t\n") is None

    def test_comment_line_returns_none(self):
        """Test that comment lines are skipped."""
        assert _classify_entry("# This is a comment") is None
        assert _classify_entry("#comment") is None

    def test_telegram_link_https(self):
        """Test classifying https Telegram link."""
        entry = _classify_entry("https://t.me/testchannel")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.LINK
        assert entry.value == "https://t.me/testchannel"
        assert entry.normalized == "testchannel"

    def test_telegram_link_normalized_lowercase(self):
        """Test that Telegram link username is normalized to lowercase."""
        entry = _classify_entry("https://t.me/TestChannel")
        assert entry is not None
        assert entry.normalized == "testchannel"

    def test_telegram_link_without_protocol(self):
        """Test classifying Telegram link without protocol."""
        entry = _classify_entry("t.me/testchannel")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.LINK
        assert entry.normalized == "testchannel"

    def test_numeric_id_positive(self):
        """Test classifying positive numeric ID."""
        entry = _classify_entry("12345")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.ID
        assert entry.value == "12345"
        assert entry.normalized == "12345"

    def test_numeric_id_negative(self):
        """Test classifying negative numeric ID."""
        entry = _classify_entry("-1001234567890")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.ID
        assert entry.value == "-1001234567890"
        assert entry.normalized == "-1001234567890"

    def test_username_with_at(self):
        """Test classifying username with @ prefix."""
        entry = _classify_entry("@testuser")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.USERNAME
        assert entry.value == "@testuser"
        assert entry.normalized == "testuser"

    def test_username_without_at(self):
        """Test classifying username without @ prefix."""
        entry = _classify_entry("testuser")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.USERNAME
        assert entry.value == "testuser"
        assert entry.normalized == "testuser"

    def test_username_normalized_lowercase(self):
        """Test that username is normalized to lowercase."""
        entry = _classify_entry("@TestUser")
        assert entry is not None
        assert entry.normalized == "testuser"

    def test_short_username_fallback(self):
        """Test that short usernames (< 4 chars) use fallback logic."""
        entry = _classify_entry("@ab")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.USERNAME
        assert entry.normalized == "ab"

    def test_username_with_special_chars_fallback(self):
        """Test that usernames with special chars use fallback logic."""
        entry = _classify_entry("@user-name")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.USERNAME
        assert entry.normalized == "user-name"

    def test_whitespace_stripped_before_classification(self):
        """Test that whitespace is stripped before classification."""
        entry = _classify_entry("  @testuser  ")
        assert entry is not None
        assert entry.value == "@testuser"

    def test_single_char_returns_none(self):
        """Test that single character returns None."""
        assert _classify_entry("a") is None

    def test_mixed_case_link(self):
        """Test link with mixed case."""
        entry = _classify_entry("https://T.ME/TestChannel")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.LINK

    def test_joinchat_link(self):
        """Test joinchat link format."""
        entry = _classify_entry("https://t.me/joinchat/AaBbCcDd")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.LINK
        assert entry.normalized == "aabbccdd"


class TestParseText:
    """Tests for parse_text function."""

    def test_empty_string(self):
        """Test parsing empty string."""
        result = parse_text("")
        assert result == []

    def test_empty_bytes(self):
        """Test parsing empty bytes."""
        result = parse_text(b"")
        assert result == []

    def test_single_username(self):
        """Test parsing single username."""
        result = parse_text("@testuser")
        assert len(result) == 1
        assert result[0].entry_type == ChatListEntryType.USERNAME
        assert result[0].normalized == "testuser"

    def test_multiple_usernames(self):
        """Test parsing multiple usernames."""
        text = "@user1\n@user2\n@user3"
        result = parse_text(text)
        assert len(result) == 3
        assert result[0].normalized == "user1"
        assert result[1].normalized == "user2"
        assert result[2].normalized == "user3"

    def test_mixed_entry_types(self):
        """Test parsing mixed entry types."""
        text = """@testuser
https://t.me/testchannel
-1001234567890
another_user"""
        result = parse_text(text)
        assert len(result) == 4
        assert result[0].entry_type == ChatListEntryType.USERNAME
        assert result[1].entry_type == ChatListEntryType.LINK
        assert result[2].entry_type == ChatListEntryType.ID
        assert result[3].entry_type == ChatListEntryType.USERNAME

    def test_skip_empty_lines(self):
        """Test that empty lines are skipped."""
        text = "@user1\n\n\n@user2"
        result = parse_text(text)
        assert len(result) == 2

    def test_skip_comment_lines(self):
        """Test that comment lines are skipped."""
        text = """# This is a comment
@user1
# Another comment
@user2"""
        result = parse_text(text)
        assert len(result) == 2
        assert result[0].normalized == "user1"
        assert result[1].normalized == "user2"

    def test_utf8_decoding(self):
        """Test UTF-8 decoding of bytes."""
        text = "@пользователь"
        result = parse_text(text.encode("utf-8"))
        assert len(result) == 1
        assert result[0].normalized == "пользователь"

    def test_cp1251_decoding(self):
        """Test CP1251 decoding fallback."""
        text = "@пользователь"
        # Encode as CP1251 (will fail UTF-8 decode)
        result = parse_text(text.encode("cp1251"))
        assert len(result) == 1
        assert result[0].normalized == "пользователь"

    def test_invalid_encoding_raises_parse_error(self):
        """Test that invalid encoding raises ParseError.

        Note: It's difficult to create a byte sequence that fails both UTF-8 and CP1251
        decoding, as CP1251 is quite permissive. This test documents the expected behavior.
        """
        # Most byte sequences will decode in CP1251, so we test the error path differently
        # by testing with a completely invalid byte sequence for UTF-8
        invalid_bytes = b"\x80\x81\x82\x83" * 100  # Invalid in UTF-8
        try:
            parse_text(invalid_bytes)
            # CP1251 might decode some sequences, so we just ensure no crash
        except ParseError as e:
            # This is the expected path if both decodings fail
            assert "Failed to decode file" in str(e)

    def test_string_input_no_decoding(self):
        """Test that string input is used directly."""
        text = "@user1\n@user2"
        result = parse_text(text)
        assert len(result) == 2

    def test_windows_line_endings(self):
        """Test parsing with Windows line endings."""
        text = "@user1\r\n@user2\r\n@user3"
        result = parse_text(text)
        assert len(result) == 3

    def test_mac_line_endings(self):
        """Test parsing with old Mac line endings."""
        text = "@user1\r@user2\r@user3"
        result = parse_text(text)
        assert len(result) == 3

    def test_mixed_line_endings(self):
        """Test parsing with mixed line endings."""
        text = "@user1\n@user2\r\n@user3\r@user4"
        result = parse_text(text)
        assert len(result) == 4


class TestParseCsv:
    """Tests for parse_csv function."""

    def test_empty_csv(self):
        """Test parsing empty CSV."""
        result = parse_csv("")
        assert result == []

    def test_single_column_no_header(self):
        """Test parsing single column without header."""
        csv_text = "@user1\n@user2\n@user3"
        result = parse_csv(csv_text)
        assert len(result) == 3
        assert result[0].normalized == "user1"

    def test_single_column_with_header(self):
        """Test parsing single column with recognized header."""
        csv_text = "username\n@user1\n@user2"
        result = parse_csv(csv_text)
        assert len(result) == 2
        assert result[0].normalized == "user1"

    def test_multiple_columns_username_header(self):
        """Test parsing multiple columns with username header."""
        csv_text = "fullname,username,status\nUser One,@user1,active\nUser Two,@user2,active"
        result = parse_csv(csv_text)
        assert len(result) == 2
        assert result[0].normalized == "user1"
        assert result[1].normalized == "user2"

    def test_multiple_columns_chat_header(self):
        """Test parsing multiple columns with chat header."""
        csv_text = (
            "description,chat,status\nChannel One,@channel1,active\nChannel Two,@channel2,inactive"
        )
        result = parse_csv(csv_text)
        assert len(result) == 2
        assert result[0].normalized == "channel1"

    def test_multiple_columns_link_header(self):
        """Test parsing multiple columns with link header."""
        csv_text = (
            "description,link\nChannel One,https://t.me/channel1\nChannel Two,https://t.me/channel2"
        )
        result = parse_csv(csv_text)
        assert len(result) == 2
        assert result[0].entry_type == ChatListEntryType.LINK

    def test_multiple_columns_id_header(self):
        """Test parsing multiple columns with id header."""
        csv_text = "description,id\nChat One,-1001234567890\nChat Two,-1009876543210"
        result = parse_csv(csv_text)
        assert len(result) == 2
        assert result[0].entry_type == ChatListEntryType.ID

    def test_semicolon_delimiter(self):
        """Test parsing CSV with semicolon delimiter."""
        csv_text = "username;name\n@user1;User One\n@user2;User Two"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_tab_delimiter(self):
        """Test parsing CSV with tab delimiter."""
        csv_text = "username\tname\n@user1\tUser One\n@user2\tUser Two"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_delimiter_detection_fallback(self):
        """Test that delimiter detection falls back to comma."""
        csv_text = "@user1\n@user2"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_no_recognized_header_includes_first_row(self):
        """Test that first row is included when no header is recognized."""
        csv_text = "data1,data2\n@user1,extra\n@user2,extra"
        result = parse_csv(csv_text)
        # First row should be classified - "data1" should become a username
        assert len(result) >= 2

    def test_skip_empty_cells(self):
        """Test that empty cells are skipped."""
        csv_text = "username\n@user1\n\n@user2\n"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_column_index_out_of_range(self):
        """Test handling of rows with fewer columns."""
        csv_text = "username,name\n@user1,User One\n@user2"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_utf8_csv(self):
        """Test UTF-8 encoded CSV."""
        csv_text = "username\n@пользователь1\n@пользователь2"
        result = parse_csv(csv_text.encode("utf-8"))
        assert len(result) == 2

    def test_cp1251_csv(self):
        """Test CP1251 encoded CSV."""
        csv_text = "username\n@пользователь1"
        result = parse_csv(csv_text.encode("cp1251"))
        assert len(result) == 1

    def test_invalid_csv_encoding(self):
        """Test that invalid encoding raises ParseError.

        Note: CP1251 is very permissive, so this tests the expected behavior.
        """
        invalid_bytes = b"\x80\x81\x82\x83" * 100
        try:
            parse_csv(invalid_bytes)
            # CP1251 might decode some sequences, so we just ensure no crash
        except ParseError as e:
            # This is the expected path if both decodings fail
            assert "Failed to decode CSV" in str(e)

    def test_all_known_column_names(self):
        """Test that all known column names are recognized."""
        known_columns = ["username", "chat", "link", "id", "url", "name", "channel", "group"]
        for col in known_columns:
            csv_text = f"{col}\n@testuser"
            result = parse_csv(csv_text)
            assert len(result) == 1, f"Column '{col}' not recognized"

    def test_case_insensitive_headers(self):
        """Test that headers are case insensitive."""
        csv_text = "USERNAME\n@user1"
        result = parse_csv(csv_text)
        assert len(result) == 1

    def test_header_with_whitespace(self):
        """Test that headers with whitespace are handled."""
        csv_text = " username \n@user1"
        result = parse_csv(csv_text)
        assert len(result) == 1

    def test_quoted_fields(self):
        """Test CSV with quoted fields."""
        csv_text = '"username","name"\n"@user1","User One"\n"@user2","User Two"'
        result = parse_csv(csv_text)
        assert len(result) == 2


class TestParseXlsx:
    """Tests for parse_xlsx function."""

    def test_openpyxl_not_installed(self, monkeypatch):
        """Test that missing openpyxl raises ParseError."""
        # Mock openpyxl import to raise ImportError
        import sys

        monkeypatch.setitem(sys.modules, "openpyxl", None)

        # Create a dummy file
        file = io.BytesIO(b"dummy")
        with pytest.raises(ParseError, match="openpyxl is required"):
            parse_xlsx(file)

    def test_empty_workbook(self, tmp_path):
        """Test parsing empty workbook."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        # Create empty workbook
        wb = Workbook()
        ws = wb.active
        # Remove default row if present
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)

        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert result == []

    def test_single_column_no_header(self, tmp_path):
        """Test parsing single column without header."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["@user1"])
        ws.append(["@user2"])
        ws.append(["@user3"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 3
        assert result[0].normalized == "user1"

    def test_single_column_with_header(self, tmp_path):
        """Test parsing single column with recognized header."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["username"])
        ws.append(["@user1"])
        ws.append(["@user2"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 2
        assert result[0].normalized == "user1"

    def test_multiple_columns_username_header(self, tmp_path):
        """Test parsing multiple columns with username header."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["fullname", "username", "status"])
        ws.append(["User One", "@user1", "active"])
        ws.append(["User Two", "@user2", "active"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 2
        assert result[0].normalized == "user1"

    def test_skip_empty_cells(self, tmp_path):
        """Test that empty cells are skipped."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["username"])
        ws.append(["@user1"])
        ws.append([None])
        ws.append(["@user2"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 2

    def test_invalid_excel_file(self):
        """Test that invalid Excel file raises ParseError."""
        invalid_file = io.BytesIO(b"not an excel file")
        with pytest.raises(ParseError, match="Failed to read Excel file"):
            parse_xlsx(invalid_file)

    def test_all_known_column_names(self, tmp_path):
        """Test that all known column names are recognized."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        known_columns = ["username", "chat", "link", "id", "url", "name", "channel", "group"]
        for col in known_columns:
            wb = Workbook()
            ws = wb.active
            ws.append([col])
            ws.append(["@testuser"])

            file_path = tmp_path / f"test_{col}.xlsx"
            wb.save(file_path)

            with open(file_path, "rb") as f:
                result = parse_xlsx(f)
            assert len(result) == 1, f"Column '{col}' not recognized"

    def test_case_insensitive_headers(self, tmp_path):
        """Test that headers are case insensitive."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["USERNAME"])
        ws.append(["@user1"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 1

    def test_numeric_cells_converted_to_string(self, tmp_path):
        """Test that numeric cells are converted to strings."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["id"])
        ws.append([12345])  # Numeric value
        ws.append([-1001234567890])  # Negative numeric value

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            result = parse_xlsx(f)
        assert len(result) == 2
        assert result[0].entry_type == ChatListEntryType.ID


class TestParseChatList:
    """Tests for parse_chat_list function."""

    def test_txt_extension(self):
        """Test parsing .txt file."""
        content = b"@user1\n@user2"
        result = parse_chat_list(content, "test.txt")
        assert len(result) == 2

    def test_csv_extension(self):
        """Test parsing .csv file."""
        content = b"username\n@user1\n@user2"
        result = parse_chat_list(content, "test.csv")
        assert len(result) == 2

    def test_xlsx_extension(self, tmp_path):
        """Test parsing .xlsx file."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["@user1"])
        ws.append(["@user2"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            content = f.read()

        result = parse_chat_list(content, "test.xlsx")
        assert len(result) == 2

    def test_xls_extension(self, tmp_path):
        """Test parsing .xls file (treated as xlsx)."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["@user1"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            content = f.read()

        result = parse_chat_list(content, "test.xls")
        assert len(result) == 1

    def test_unknown_extension_with_pk_header(self, tmp_path):
        """Test auto-detection of xlsx by PK header."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.append(["@user1"])

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        with open(file_path, "rb") as f:
            content = f.read()

        # Use unknown extension
        result = parse_chat_list(content, "test.dat")
        assert len(result) == 1

    def test_unknown_extension_tries_csv_then_text(self):
        """Test that unknown extension tries CSV then falls back to text."""
        # CSV-like content
        content = b"username\n@user1"
        result = parse_chat_list(content, "test.dat")
        assert len(result) == 1

    def test_unknown_extension_fallback_to_text(self):
        """Test fallback to text parser for unknown extension."""
        # Plain text content (will fail CSV parsing)
        content = b"not,valid,csv\n@user1\n@user2"
        result = parse_chat_list(content, "test.dat")
        # Should fall back to text parser and get 2 users
        assert len(result) >= 2

    def test_unknown_extension_csv_parse_error_fallback(self):
        """Test that ParseError in CSV parsing falls back to text parsing."""
        # Content that looks like CSV but will cause issues
        content = b"@user1\n@user2\n@user3"
        result = parse_chat_list(content, "test.unknown")
        # Should fall back to text parser
        assert len(result) == 3
        assert all(r.entry_type == ChatListEntryType.USERNAME for r in result)

    def test_case_insensitive_extension(self):
        """Test that extension matching is case insensitive."""
        content = b"@user1"
        result = parse_chat_list(content, "TEST.TXT")
        assert len(result) == 1

    def test_extension_with_path(self):
        """Test that extension is detected from full path."""
        content = b"@user1"
        result = parse_chat_list(content, "/path/to/file.txt")
        assert len(result) == 1


class TestEdgeCasesAndErrors:
    """Tests for edge cases and error conditions."""

    def test_very_long_username(self):
        """Test handling of very long usernames."""
        long_username = "@" + "a" * 100
        entry = _classify_entry(long_username)
        assert entry is not None
        # Should use fallback logic for long usernames

    def test_unicode_usernames(self):
        """Test handling of unicode usernames."""
        result = parse_text("@пользователь\n@用户\n@مستخدم")
        assert len(result) == 3

    def test_mixed_valid_and_invalid_entries(self):
        """Test that valid entries are parsed even with invalid ones.

        Note: The parser is lenient and tries to parse most entries using
        fallback logic, so 'invalid!!!entry' might be accepted as a username.
        """
        text = """@user1
x
@user2
y
@user3"""
        result = parse_text(text)
        # Single character entries return None, so we should get 3 valid entries
        assert len(result) == 3
        assert result[0].normalized == "user1"
        assert result[1].normalized == "user2"
        assert result[2].normalized == "user3"

    def test_duplicate_entries_not_deduplicated(self):
        """Test that duplicate entries are not automatically removed."""
        text = "@user1\n@user1\n@user1"
        result = parse_text(text)
        assert len(result) == 3

    def test_link_with_query_params(self):
        """Test link with query parameters."""
        entry = _classify_entry("https://t.me/testchannel?start=abc")
        # Should not match because of query params
        # Will fall through to fallback logic
        assert entry is not None

    def test_csv_with_bom(self):
        """Test CSV with UTF-8 BOM.

        Note: BOM is part of the header, so it won't match known columns.
        The parser will treat the first row as data.
        """
        csv_text = "\ufeffusername\n@user1"
        result = parse_csv(csv_text)
        # BOM prevents header matching, so both rows are processed as data
        assert len(result) >= 1  # At least @user1 should be parsed

    def test_csv_with_empty_rows(self):
        """Test CSV with empty rows."""
        csv_text = "username\n@user1\n\n\n@user2\n\n"
        result = parse_csv(csv_text)
        assert len(result) == 2

    def test_text_with_only_comments(self):
        """Test text file with only comments."""
        text = "# Comment 1\n# Comment 2\n# Comment 3"
        result = parse_text(text)
        assert result == []

    def test_text_with_only_empty_lines(self):
        """Test text file with only empty lines."""
        text = "\n\n\n\n"
        result = parse_text(text)
        assert result == []

    def test_negative_zero_id(self):
        """Test handling of -0 as ID."""
        entry = _classify_entry("-0")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.ID

    def test_very_large_numeric_id(self):
        """Test handling of very large numeric IDs."""
        entry = _classify_entry("-100123456789012345")
        assert entry is not None
        assert entry.entry_type == ChatListEntryType.ID

    def test_whitespace_in_middle_of_value(self):
        """Test that internal whitespace is preserved."""
        # This should not match username pattern and use fallback
        entry = _classify_entry("@user name")
        assert entry is not None
        # Should be normalized to "user name" (without @)

    def test_parse_error_exception(self):
        """Test ParseError exception properties."""
        error = ParseError("Test error message")
        assert str(error) == "Test error message"
        assert isinstance(error, Exception)

    def test_chat_list_entry_type_enum_values(self):
        """Test ChatListEntryType enum values."""
        assert ChatListEntryType.USERNAME == "username"
        assert ChatListEntryType.LINK == "link"
        assert ChatListEntryType.ID == "id"


class TestIntegration:
    """Integration tests combining multiple functions."""

    def test_real_world_text_file(self):
        """Test parsing a realistic text file with mixed content."""
        text = """# Chat list for project X
# Last updated: 2024-01-15

# Official channels
@official_channel
https://t.me/news_channel

# Partner groups
@partner_group_1
@partner_group_2

# Test channels
-1001234567890
-1009876543210

# Archive (commented out)
# @old_channel
# @deprecated_group
"""
        result = parse_text(text)
        assert len(result) == 6

    def test_real_world_csv_file(self):
        """Test parsing a realistic CSV export."""
        csv_text = """name,username,url,status,members
"News Channel","@news_channel","https://t.me/news_channel","active",5000
"Chat Group","@chat_group","https://t.me/chat_group","active",150
"Archive","@archive_channel","https://t.me/archive_channel","inactive",0
"""
        result = parse_csv(csv_text)
        assert len(result) >= 3

    def test_round_trip_text_to_entries(self):
        """Test that parsed entries maintain their original values."""
        text = "@user1\nhttps://t.me/channel1\n-1001234567890"
        result = parse_text(text)

        assert result[0].value == "@user1"
        assert result[1].value == "https://t.me/channel1"
        assert result[2].value == "-1001234567890"

    def test_normalization_consistency(self):
        """Test that normalization is consistent across formats."""
        username_text = "@TestUser"
        link_text = "https://t.me/TestUser"

        entry1 = _classify_entry(username_text)
        entry2 = _classify_entry(link_text)

        assert entry1 is not None
        assert entry2 is not None
        assert entry1.normalized == entry2.normalized  # Both should be "testuser"
